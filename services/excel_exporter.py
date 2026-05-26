import io
from datetime import datetime
from services.ice_calculator import IceCalculator, TAX_DB

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


class ExcelExporter:
    
    @staticmethod
    def crear_estilos(wb):
        return {
            'h_fill': PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid'),
            'h_font': Font(name='Arial', size=9, bold=True, color='FFFFFF'),
            'total_fill': PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid'),
            'total_font': Font(name='Arial', size=10, bold=True, color='FFFFFF'),
            'nfont': Font(name='Arial', size=9),
            'bfont': Font(name='Arial', size=9, bold=True),
            'title_font': Font(name='Arial', size=14, bold=True, color='1a5276'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }
    
    @staticmethod
    def exportar_declaracion(facturas):
        """Exporta Excel con el detalle de facturas (como SRI-XML.py)"""
        if not EXCEL_DISPONIBLE:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Declaracion"
        s = ExcelExporter.crear_estilos(wb)
        
        encabezados = ['Fecha', 'N Factura', 'RUC Emisor', 'Cliente', 'Total',
                       'Base ICE', 'ICE', 'Base IVA', 'IVA 15%', 'Importe Total']
        
        for j, enc in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=j, value=enc)
            cell.fill = s['h_fill']
            cell.font = s['h_font']
            cell.border = s['border']
            cell.alignment = Alignment(horizontal='center')
        
        fila = 2
        for f in facturas:
            vals = [
                f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else '',
                f.numero_factura or '',
                f.ruc_emisor or '',
                f.razon_social_comprador[:40] if f.razon_social_comprador else '',
                float(f.importe_total or 0),
                float(f.base_ice or 0),
                float(f.valor_ice or 0),
                float(f.base_iva or 0),
                float(f.valor_iva or 0),
                float(f.importe_total or 0)
            ]
            for j, val in enumerate(vals, 1):
                cell = ws.cell(row=fila, column=j, value=val)
                cell.font = s['nfont']
                cell.border = s['border']
                if j >= 5:
                    cell.number_format = '#,##0.00'
            fila += 1
        
        # Totales
        ult = fila
        for j in range(1, 11):
            cell = ws.cell(row=ult, column=j)
            cell.fill = s['total_fill']
            cell.font = s['total_font']
            cell.border = s['border']
        
        ws.cell(row=ult, column=1, value="TOTALES")
        for col in [5, 6, 7, 8, 9, 10]:
            letra = get_column_letter(col)
            ws.cell(row=ult, column=col, value=f"=SUM({letra}2:{letra}{ult-1})")
            ws.cell(row=ult, column=col).number_format = '#,##0.00'
        
        for j, w in enumerate([12, 15, 15, 30, 12, 12, 12, 12, 12, 12], 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def exportar_auditoria(datos_producto, anios_seleccionados):
        """Exporta Excel de auditoria ICE multi-ano (como ICEcapsa.py)"""
        if not EXCEL_DISPONIBLE:
            return None
        
        wb = Workbook()
        s = ExcelExporter.crear_estilos(wb)
        
        # Hoja 1: Auditoria
        ws = wb.active
        ws.title = "Auditoria ICE"
        
        ws.cell(row=1, column=1, value="AUDITORIA ICE - CALCULO POR PRODUCTO").font = s['title_font']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        
        enc = ['Ano', 'Tarifa Esp.', 'Umbral AdV', 'IVA', 'ICE Esp. Unit.',
               'ICE AdV. Unit.', 'ICE Total Unit.', 'Base IVA', 'IVA Total', 'PVP Final']
        
        for j, e in enumerate(enc, 1):
            cell = ws.cell(row=3, column=j, value=e)
            cell.fill = s['h_fill']
            cell.font = s['h_font']
            cell.border = s['border']
        
        fila = 4
        for anio in anios_seleccionados:
            res = IceCalculator.calcular_liquidacion_completa(datos_producto, anio)
            vals = [
                anio, res['tarifa_especifica'], res['umbral_advalorem'],
                f"{int(res['iva_tasa']*100)}%", res['ice_especifico_unitario'],
                res['ice_advalorem_unitario'], res['ice_total_unitario'],
                res['base_iva'], res['iva_total'], res['pvp']
            ]
            for j, val in enumerate(vals, 1):
                cell = ws.cell(row=fila, column=j, value=val)
                cell.font = s['nfont']
                cell.border = s['border']
                if j >= 4:
                    cell.number_format = '#,##0.0000' if j <= 7 else '#,##0.00'
            fila += 1
        
        for j in range(1, 11):
            ws.column_dimensions[get_column_letter(j)].width = 18
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output